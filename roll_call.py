#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
点名系统 - Random Student Picker
圆形排列名字，灯光旋转5-6秒后随机停在某学生上，记录回答并评分，导出到Excel。
功能：TTS语音播报 | 考勤追踪 | 快捷键 | 主题切换 | 批量导入导出 | 搜索列表
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from tkinter.font import families
import math
import random
import time
from datetime import datetime
from pathlib import Path
import json
import sys
import csv
import os
import threading

# ---- 可选依赖 ----
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    import pyttsx3
    TTS_AVAILABLE = True
except ImportError:
    TTS_AVAILABLE = False
    print("提示: 未安装 pyttsx3，语音功能不可用。运行: pip install pyttsx3")

# ============================================================
# 配置
# ============================================================
APP_DIR = Path(__file__).parent
CONFIG_FILE = APP_DIR / "config.json"
EXCEL_FILE = APP_DIR / "点名记录.xlsx"
ATTENDANCE_FILE = APP_DIR / "attendance.json"
HISTORY_FILE = APP_DIR / "history.json"
DEFAULT_STUDENTS = [
    "张三", "李四", "王五", "赵六", "钱七",
    "孙八", "周九", "吴十", "郑十一", "王小明",
    "李小红", "张小刚", "刘小丽", "陈小华", "杨小芳",
    "黄大伟", "林小红", "何小军", "马小燕", "罗小强",
]

# ============================================================
# 主题定义
# ============================================================
DARK_THEME = {
    "name": "dark",
    "bg": "#1a1a2e",
    "circle_bg": "#16213e",
    "text": "#e0e0e0",
    "highlight": "#f7b731",
    "highlight_glow": "#fed330",
    "selected": "#fc5c65",
    "accent": "#45aaf2",
    "button": "#2d98da",
    "button_hover": "#45aaf2",
    "panel_bg": "#0f3460",
    "title_bg": "#0f3460",
    "title_fg": "#f7b731",
    "info_fg": "#636e72",
    "score_colors": ["#fc5c65", "#fd9644", "#fed330", "#26de81", "#20bf6b"],
    "input_bg": "#1e3a5f",
    "input_fg": "#dfe6e9",
    "list_bg": "#1e3a5f",
    "list_fg": "#dfe6e9",
    "tab_bg": "#0f3460",
    "tab_fg": "#dfe6e9",
    "tab_selected_bg": "#16213e",
    "button_text": "white",
    "stats_label_bg": "#1e3a5f",
    "stats_label_fg": "#dfe6e9",
    "history_bg": "#1e3a5f",
    "history_fg": "#dfe6e9",
    "circle_dot": "#636e72",
    "circle_dot_outline": "#b2bec3",
    "circle_name": "#b2bec3",
    "circle_track": "#2d3436",
    "btn_manage": "#6c5ce7",
    "btn_manage_hover": "#a29bfe",
    "btn_view": "#00b894",
    "btn_view_hover": "#55efc4",
    "btn_export_atd": "#e17055",
    "btn_export_atd_hover": "#fab1a0",
    "btn_cancel": "#636e72",
    "btn_cancel_hover": "#b2bec3",
    "btn_save": "#20bf6b",
    "btn_save_hover": "#26de81",
    "btn_danger": "#e74c3c",
    "btn_danger_hover": "#ff7675",
    "btn_warning": "#f39c12",
    "btn_warning_hover": "#fdcb6e",
}

LIGHT_THEME = {
    "name": "light",
    "bg": "#f0f0f0",
    "circle_bg": "#ffffff",
    "text": "#2d3436",
    "highlight": "#f7b731",
    "highlight_glow": "#fed330",
    "selected": "#fc5c65",
    "accent": "#0984e3",
    "button": "#0984e3",
    "button_hover": "#74b9ff",
    "panel_bg": "#ffffff",
    "title_bg": "#dfe6e9",
    "title_fg": "#2d3436",
    "info_fg": "#636e72",
    "score_colors": ["#d63031", "#e17055", "#fdcb6e", "#00b894", "#0984e3"],
    "input_bg": "#ffffff",
    "input_fg": "#2d3436",
    "list_bg": "#ffffff",
    "list_fg": "#2d3436",
    "tab_bg": "#dfe6e9",
    "tab_fg": "#2d3436",
    "tab_selected_bg": "#ffffff",
    "button_text": "white",
    "stats_label_bg": "#ffffff",
    "stats_label_fg": "#2d3436",
    "history_bg": "#ffffff",
    "history_fg": "#2d3436",
    "circle_dot": "#b2bec3",
    "circle_dot_outline": "#636e72",
    "circle_name": "#636e72",
    "circle_track": "#dfe6e9",
    "btn_manage": "#6c5ce7",
    "btn_manage_hover": "#a29bfe",
    "btn_view": "#00b894",
    "btn_view_hover": "#55efc4",
    "btn_export_atd": "#e17055",
    "btn_export_atd_hover": "#fab1a0",
    "btn_cancel": "#b2bec3",
    "btn_cancel_hover": "#636e72",
    "btn_save": "#00b894",
    "btn_save_hover": "#55efc4",
    "btn_danger": "#d63031",
    "btn_danger_hover": "#ff7675",
    "btn_warning": "#e17055",
    "btn_warning_hover": "#fab1a0",
}


def get_best_chinese_font():
    """检测系统可用的中文字体，返回最佳选择"""
    available = set(families(root=None))
    candidates = [
        "Microsoft YaHei",
        "Microsoft JhengHei",
        "SimHei",
        "KaiTi",
        "SimSun",
        "DengXian",
        "FangSong",
        "STXihei",
        "STSong",
    ]
    for font in candidates:
        if font in available:
            return font
    return "TkDefaultFont"


# 延迟初始化
FONT_FAMILY = None


class RollCallApp:
    def __init__(self, root):
        self.root = root
        global FONT_FAMILY
        FONT_FAMILY = get_best_chinese_font()
        self.root.title("🎯 课堂点名系统")
        self.root.geometry("1100x720")
        self.root.configure(bg=DARK_THEME["bg"])
        self.root.minsize(1000, 650)

        # ---- 主题 ----
        self.theme = DARK_THEME.copy()
        self._dark_mode = True

        # ---- TTS 语音引擎 ----
        self.tts_engine = None
        if TTS_AVAILABLE:
            self.init_tts_engine()

        # ---- 状态变量 ----
        self.students = self.load_students()
        self.current_index = 0
        self.is_spinning = False
        self.spin_speed = 0.05
        self.spin_start_time = 0
        self.spin_duration = 5.5
        self.selected_student = None
        self.records = []
        self.call_cooldown_ms = 600  # 防快速点击冷却

        # ---- 考勤数据 ----
        self.attendance = self.load_attendance()

        # ---- 点名历史 ----
        self.call_history = self.load_history()

        # ---- 构建UI ----
        self.setup_ui()
        self.setup_shortcuts()
        self.draw_circle()

        # 窗口居中
        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        self.root.geometry(f"+{x}+{y}")

        # 绑定事件
        self.root.bind("<Escape>", lambda e: self.root.quit())
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ================================================================
    #  TTS 语音
    # ================================================================
    def init_tts_engine(self):
        """初始化语音引擎（改进的中文检测）"""
        try:
            self.tts_engine = pyttsx3.init()
            self.tts_engine.setProperty('rate', 180)
            self.tts_engine.setProperty('volume', 0.9)
            voices = self.tts_engine.getProperty('voices')
            for voice in voices:
                if 'zh' in voice.id.lower() or 'chinese' in voice.name.lower():
                    self.tts_engine.setProperty('voice', voice.id)
                    break
        except Exception as e:
            print(f"语音引擎初始化失败: {e}")
            self.tts_engine = None

    def speak_name(self, name):
        """播报学生姓名（非阻塞）"""
        if not self.tts_engine:
            return

        def _speak():
            try:
                self.tts_engine.say(name)
                self.tts_engine.runAndWait()
            except Exception:
                pass

        threading.Thread(target=_speak, daemon=True).start()

    # ================================================================
    #  数据读写
    # ================================================================
    def load_students(self):
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("students", DEFAULT_STUDENTS)
            except Exception:
                return DEFAULT_STUDENTS[:]
        return DEFAULT_STUDENTS[:]

    def save_students(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({"students": self.students}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存配置失败: {e}")

    def load_attendance(self):
        if ATTENDANCE_FILE.exists():
            try:
                with open(ATTENDANCE_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def save_attendance(self):
        try:
            with open(ATTENDANCE_FILE, "w", encoding="utf-8") as f:
                json.dump(self.attendance, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存考勤失败: {e}")

    def load_history(self):
        if HISTORY_FILE.exists():
            try:
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return []
        return []

    def save_history(self):
        try:
            with open(HISTORY_FILE, "w", encoding="utf-8") as f:
                json.dump(self.call_history[-100:], f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存历史失败: {e}")

    def add_history(self, message):
        """添加一条历史记录"""
        self.call_history.append(message)
        if len(self.call_history) > 100:
            self.call_history = self.call_history[-100:]
        self.refresh_history_tab()
        self.save_history()

    # ---- Excel 记录 ----
    def save_to_excel(self, name, answer, score, timestamp):
        try:
            if EXCEL_FILE.exists():
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "点名记录"
                ws.append(["序号", "姓名", "回答内容", "评分(1-10)", "时间"])
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="2d98da", end_color="2d98da", fill_type="solid")
                for col in range(1, 6):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

            next_row = ws.max_row + 1
            seq = next_row - 1
            ws.append([seq, name, answer, score, timestamp])

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 22

            for col in ['A', 'B', 'D', 'E']:
                ws.cell(row=next_row, column=ord(col) - ord('A') + 1).alignment = Alignment(horizontal="center")

            wb.save(EXCEL_FILE)
            return True
        except Exception as e:
            print(f"保存Excel失败: {e}")
            return False

    # ================================================================
    #  键盘快捷键
    # ================================================================
    def setup_shortcuts(self):
        self.root.bind('<F5>', lambda e: self.start_spin())
        self.root.bind('<Control-a>', lambda e: self.shortcut_add_student())
        self.root.bind('<Control-A>', lambda e: self.shortcut_add_student())
        self.root.bind('<Delete>', lambda e: self.shortcut_delete_student())
        self.root.bind('<Return>', lambda e: self.shortcut_call_selected())

    def shortcut_add_student(self):
        self.open_add_student_dialog()
        return "break"

    def shortcut_delete_student(self):
        self.delete_selected_student()
        return "break"

    def shortcut_call_selected(self):
        sel = self.student_listbox.curselection()
        if sel:
            name = self.student_listbox.get(sel[0])
            if name in self.students:
                self.selected_student = name
                self.draw_circle()
                self.info_label.config(text=f"选中: {name}  |  共 {len(self.students)} 名学生")
                self.root.after(200, self.show_scoring_dialog)

    # ================================================================
    #  主题切换
    # ================================================================
    def toggle_theme(self):
        if self._dark_mode:
            self._dark_mode = False
            self.theme = LIGHT_THEME.copy()
            self.theme_btn.config(text="🌙 暗色")
        else:
            self._dark_mode = True
            self.theme = DARK_THEME.copy()
            self.theme_btn.config(text="☀️ 亮色")

        self.apply_theme()

    def apply_theme(self):
        """应用当前主题到所有组件"""
        t = self.theme
        self.root.configure(bg=t["bg"])
        self.canvas.configure(bg=t["circle_bg"])
        self.title_frame.configure(bg=t["title_bg"])
        self.title_label.configure(fg=t["title_fg"], bg=t["title_bg"])
        self.info_label.configure(fg=t["info_fg"], bg=t["bg"])
        self.left_container.configure(bg=t["bg"])
        self.bottom_frame.configure(bg=t["bg"])
        self.btn_frame.configure(bg=t["bg"])

        # 按钮
        self.spin_btn.configure(bg=t["button"], fg=t["button_text"],
                                activebackground=t["button_hover"])
        self.manage_btn.configure(bg=t["btn_manage"], fg=t["button_text"],
                                  activebackground=t["btn_manage_hover"])
        self.view_btn.configure(bg=t["btn_view"], fg=t["button_text"],
                                activebackground=t["btn_view_hover"])
        if hasattr(self, 'speak_btn'):
            self.speak_btn.configure(bg=t["accent"], fg=t["button_text"],
                                     activebackground=t["button_hover"])

        # 右侧面板
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background=t["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=t["tab_bg"], foreground=t["tab_fg"],
                        padding=[10, 4], font=(FONT_FAMILY, 10))
        style.map("TNotebook.Tab",
                  background=[("selected", t["tab_selected_bg"])],
                  foreground=[("selected", t["accent"])])
        style.configure("TFrame", background=t["bg"])

        self.right_panel.configure(bg=t["bg"])
        self.tab_students.configure(bg=t["bg"])
        self.tab_attendance.configure(bg=t["bg"])
        self.tab_history.configure(bg=t["bg"])

        # 学生名单 Tab
        self.student_search_frame.configure(bg=t["bg"])
        self.student_listbox.configure(bg=t["list_bg"], fg=t["list_fg"])
        self.student_btn_frame.configure(bg=t["bg"])

        # 考勤 Tab
        if hasattr(self, 'stats_frame'):
            self.stats_frame.configure(bg=t["bg"])
        if hasattr(self, 'stats_label'):
            self.stats_label.configure(bg=t["stats_label_bg"], fg=t["stats_label_fg"])

        # 历史 Tab
        if hasattr(self, 'history_text'):
            self.history_text.configure(bg=t["history_bg"], fg=t["history_fg"])

        self.refresh_stats()
        self.draw_circle()

    # ================================================================
    #  UI 构建
    # ================================================================
    def setup_ui(self):
        t = self.theme

        # ---- 顶部标题栏 ----
        self.title_frame = tk.Frame(self.root, bg=t["title_bg"], height=46)
        self.title_frame.pack(fill=tk.X, side=tk.TOP)
        self.title_frame.pack_propagate(False)

        self.title_label = tk.Label(
            self.title_frame, text="🎯 课堂点名系统",
            font=(FONT_FAMILY, 18, "bold"),
            fg=t["title_fg"], bg=t["title_bg"]
        )
        self.title_label.pack(side=tk.LEFT, padx=(20, 0), pady=8)

        self.theme_btn = tk.Button(
            self.title_frame, text="☀️ 亮色", font=(FONT_FAMILY, 9),
            bg=t["btn_cancel"], fg="white",
            activebackground=t["btn_cancel_hover"],
            relief=tk.FLAT, padx=10, pady=3, cursor="hand2",
            command=self.toggle_theme
        )
        self.theme_btn.pack(side=tk.RIGHT, padx=20, pady=8)

        # ---- 主容器（左右分栏） ----
        main_container = tk.Frame(self.root, bg=t["bg"])
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 5))

        # ==== 左侧：画布 + 控制按钮 ====
        self.left_container = tk.Frame(main_container, bg=t["bg"])
        self.left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(
            self.left_container, bg=t["circle_bg"], highlightthickness=0,
            width=650, height=420
        )
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=(0, 5), pady=(0, 5))
        # Canvas resize 时重新绘制
        self.canvas.bind("<Configure>", lambda e: self.draw_circle())

        # 底部控制
        self.bottom_frame = tk.Frame(self.left_container, bg=t["bg"])
        self.bottom_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(5, 5))

        self.btn_frame = tk.Frame(self.bottom_frame, bg=t["bg"])
        self.btn_frame.pack(pady=(5, 8))

        self.spin_btn = tk.Button(
            self.btn_frame, text="🎰 开始点名 (F5)", font=(FONT_FAMILY, 13, "bold"),
            bg=t["button"], fg=t["button_text"],
            activebackground=t["button_hover"],
            activeforeground="white", relief=tk.FLAT,
            padx=24, pady=7, cursor="hand2",
            command=self.start_spin
        )
        self.spin_btn.pack(side=tk.LEFT, padx=6)

        self.speak_btn = tk.Button(
            self.btn_frame, text="🔊 重新播报", font=(FONT_FAMILY, 11),
            bg=t["accent"], fg=t["button_text"],
            activebackground=t["button_hover"],
            activeforeground="white", relief=tk.FLAT,
            padx=14, pady=7, cursor="hand2",
            command=self.re_speak
        )
        self.speak_btn.pack(side=tk.LEFT, padx=6)

        self.manage_btn = tk.Button(
            self.btn_frame, text="📋 管理名单", font=(FONT_FAMILY, 11),
            bg=t["btn_manage"], fg=t["button_text"],
            activebackground=t["btn_manage_hover"],
            activeforeground="white", relief=tk.FLAT,
            padx=14, pady=7, cursor="hand2",
            command=self.open_student_manager
        )
        self.manage_btn.pack(side=tk.LEFT, padx=6)

        self.view_btn = tk.Button(
            self.btn_frame, text="📊 查看记录", font=(FONT_FAMILY, 11),
            bg=t["btn_view"], fg=t["button_text"],
            activebackground=t["btn_view_hover"],
            activeforeground="white", relief=tk.FLAT,
            padx=14, pady=7, cursor="hand2",
            command=self.open_records
        )
        self.view_btn.pack(side=tk.LEFT, padx=6)

        self.info_label = tk.Label(
            self.bottom_frame,
            text=f"共 {len(self.students)} 名学生  |  就绪  |  F5点名  Enter直接点名选中学生",
            font=(FONT_FAMILY, 9), fg=t["info_fg"], bg=t["bg"]
        )
        self.info_label.pack(pady=(0, 2))

        # ==== 右侧面板（Notebook 选项卡） ====
        self.right_panel = tk.Frame(main_container, bg=t["bg"], width=320)
        self.right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(5, 0))
        self.right_panel.pack_propagate(False)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background=t["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=t["tab_bg"], foreground=t["tab_fg"],
                        padding=[10, 4], font=(FONT_FAMILY, 10))
        style.map("TNotebook.Tab",
                  background=[("selected", t["tab_selected_bg"])],
                  foreground=[("selected", t["accent"])])
        style.configure("TFrame", background=t["bg"])

        self.notebook = ttk.Notebook(self.right_panel)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # -- Tab 1: 学生名单 --
        self.tab_students = tk.Frame(self.notebook, bg=t["bg"])
        self.notebook.add(self.tab_students, text="  📋 学生名单  ")
        self.build_student_tab()

        # -- Tab 2: 考勤统计 --
        self.tab_attendance = tk.Frame(self.notebook, bg=t["bg"])
        self.notebook.add(self.tab_attendance, text="  📊 考勤统计  ")
        self.build_attendance_tab()

        # -- Tab 3: 点名历史 --
        self.tab_history = tk.Frame(self.notebook, bg=t["bg"])
        self.notebook.add(self.tab_history, text="  📜 点名历史  ")
        self.build_history_tab()

    def build_student_tab(self):
        t = self.theme

        # 搜索框
        self.student_search_frame = tk.Frame(self.tab_students, bg=t["bg"])
        self.student_search_frame.pack(fill=tk.X, padx=8, pady=(10, 5))

        tk.Label(self.student_search_frame, text="🔍", font=(FONT_FAMILY, 10),
                 bg=t["bg"], fg=t["text"]).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *a: self.search_students())
        search_entry = tk.Entry(
            self.student_search_frame, textvariable=self.search_var,
            font=(FONT_FAMILY, 10), width=18,
            bg=t["input_bg"], fg=t["input_fg"],
            insertbackground=t["text"],
            relief=tk.FLAT, highlightthickness=1,
            highlightbackground=t["accent"]
        )
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # 学生列表
        list_container = tk.Frame(self.tab_students, bg=t["bg"])
        list_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)

        scrollbar = tk.Scrollbar(list_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.student_listbox = tk.Listbox(
            list_container, font=(FONT_FAMILY, 10),
            yscrollcommand=scrollbar.set,
            selectmode=tk.SINGLE, height=14,
            bg=t["list_bg"], fg=t["list_fg"],
            selectbackground=t["accent"],
            selectforeground="white",
            relief=tk.FLAT, highlightthickness=1,
            highlightbackground=t["accent"]
        )
        self.student_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.student_listbox.yview)
        self.student_listbox.bind('<Double-Button-1>', lambda e: self.call_from_list())

        self.refresh_student_list()

        # 操作按钮
        self.student_btn_frame = tk.Frame(self.tab_students, bg=t["bg"])
        self.student_btn_frame.pack(fill=tk.X, padx=8, pady=(5, 10))

        tk.Button(
            self.student_btn_frame, text="➕", font=(FONT_FAMILY, 10),
            bg=t["button"], fg="white", relief=tk.FLAT,
            padx=8, pady=3, cursor="hand2",
            command=self.open_add_student_dialog
        ).pack(side=tk.LEFT, padx=2)

        tk.Button(
            self.student_btn_frame, text="➖", font=(FONT_FAMILY, 10),
            bg=t["btn_danger"], fg="white", relief=tk.FLAT,
            padx=8, pady=3, cursor="hand2",
            command=self.delete_selected_student
        ).pack(side=tk.LEFT, padx=2)

        tk.Button(
            self.student_btn_frame, text="📥 导入", font=(FONT_FAMILY, 9),
            bg=t["btn_manage"], fg="white", relief=tk.FLAT,
            padx=10, pady=3, cursor="hand2",
            command=self.batch_import_students
        ).pack(side=tk.LEFT, padx=2)

        tk.Button(
            self.student_btn_frame, text="📋 导出", font=(FONT_FAMILY, 9),
            bg=t["btn_view"], fg="white", relief=tk.FLAT,
            padx=10, pady=3, cursor="hand2",
            command=self.export_student_list
        ).pack(side=tk.LEFT, padx=2)

    def build_attendance_tab(self):
        t = self.theme

        # 标题
        tk.Label(
            self.tab_attendance, text="📊 今日考勤统计",
            font=(FONT_FAMILY, 13, "bold"),
            fg=t["accent"], bg=t["bg"]
        ).pack(pady=(15, 10))

        self.stats_frame = tk.Frame(self.tab_attendance, bg=t["stats_label_bg"],
                                     relief=tk.FLAT, highlightthickness=1,
                                     highlightbackground=t["accent"])
        self.stats_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.stats_label = tk.Label(
            self.stats_frame, text="",
            font=(FONT_FAMILY, 11),
            bg=t["stats_label_bg"], fg=t["stats_label_fg"],
            justify=tk.LEFT, padx=15, pady=10
        )
        self.stats_label.pack(fill=tk.BOTH, expand=True)

        # 导出考勤按钮
        btn_frame = tk.Frame(self.tab_attendance, bg=t["bg"])
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 15))

        tk.Button(
            btn_frame, text="📥 导出考勤报表", font=(FONT_FAMILY, 11),
            bg=t["btn_export_atd"], fg="white",
            activebackground=t["btn_export_atd_hover"],
            relief=tk.FLAT, padx=14, pady=5, cursor="hand2",
            command=self.export_attendance_report
        ).pack(fill=tk.X)

        self.refresh_stats()

    def build_history_tab(self):
        t = self.theme

        tk.Label(
            self.tab_history, text="📜 最近点名记录",
            font=(FONT_FAMILY, 13, "bold"),
            fg=t["accent"], bg=t["bg"]
        ).pack(pady=(15, 10))

        history_container = tk.Frame(self.tab_history, bg=t["bg"])
        history_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 10))

        scrollbar = tk.Scrollbar(history_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.history_text = tk.Text(
            history_container, font=(FONT_FAMILY, 9),
            yscrollcommand=scrollbar.set,
            wrap=tk.WORD, state=tk.DISABLED,
            bg=t["history_bg"], fg=t["history_fg"],
            relief=tk.FLAT, highlightthickness=1,
            highlightbackground=t["accent"],
            padx=8, pady=6, height=14
        )
        self.history_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.history_text.yview)

        self.refresh_history_tab()

    # ================================================================
    #  学生列表操作
    # ================================================================
    def refresh_student_list(self, filter_text=""):
        """刷新学生列表框"""
        self.student_listbox.delete(0, tk.END)
        for name in sorted(self.students):
            if filter_text.lower() in name.lower():
                self.student_listbox.insert(tk.END, name)

    def search_students(self):
        self.refresh_student_list(self.search_var.get().strip())

    def open_add_student_dialog(self):
        """弹出添加学生对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加学生")
        dialog.geometry("350x180")
        dialog.configure(bg=self.theme["panel_bg"])
        dialog.transient(self.root)
        dialog.grab_set()

        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() - w) // 2
        y = (dialog.winfo_screenheight() - h) // 2
        dialog.geometry(f"+{x}+{y}")

        tk.Label(dialog, text="学生姓名:", font=(FONT_FAMILY, 11),
                 bg=self.theme["panel_bg"], fg=self.theme["text"]).pack(pady=15)
        entry = tk.Entry(dialog, font=(FONT_FAMILY, 11), width=20,
                         bg=self.theme["input_bg"], fg=self.theme["input_fg"])
        entry.pack(pady=5)
        entry.focus_set()

        def add():
            name = entry.get().strip()
            if name:
                if name not in self.students:
                    self.students.append(name)
                    self.students.sort()
                    self.save_students()
                    self.refresh_student_list()
                    self.draw_circle()
                    self.info_label.config(text=f"已添加: {name}  |  共 {len(self.students)} 名学生")
                    dialog.destroy()
                else:
                    messagebox.showwarning("警告", "学生已存在！", parent=dialog)
            else:
                messagebox.showwarning("警告", "请输入学生姓名！", parent=dialog)

        tk.Button(
            dialog, text="添加", command=add,
            bg=self.theme["button"], fg="white",
            width=10, font=(FONT_FAMILY, 10), cursor="hand2"
        ).pack(pady=10)
        entry.bind('<Return>', lambda e: add())

    def delete_selected_student(self):
        """删除选中的学生"""
        sel = self.student_listbox.curselection()
        if not sel:
            messagebox.showwarning("警告", "请先在右侧列表中选择要删除的学生！")
            return

        name = self.student_listbox.get(sel[0])
        if messagebox.askyesno("确认删除", f"确定要删除学生「{name}」吗？\n该学生的考勤记录也会被删除。"):
            self.students.remove(name)
            if name in self.attendance:
                del self.attendance[name]
            self.save_students()
            self.save_attendance()
            self.refresh_student_list()
            self.draw_circle()
            self.refresh_stats()
            self.info_label.config(text=f"已删除: {name}  |  共 {len(self.students)} 名学生")

            if self.selected_student == name:
                self.selected_student = None

    def call_from_list(self):
        """双击列表项直接点名"""
        sel = self.student_listbox.curselection()
        if sel:
            name = self.student_listbox.get(sel[0])
            if name in self.students:
                self.selected_student = name
                self.draw_circle()
                self.info_label.config(text=f"选中: {name}  |  共 {len(self.students)} 名学生")
                self.show_scoring_dialog()

    def batch_import_students(self):
        """批量导入学生名单"""
        file_path = filedialog.askopenfilename(
            title="导入学生名单",
            filetypes=[("文本文件", "*.txt"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    new_names = [row[0].strip() for row in reader if row and row[0].strip()]
            else:
                with open(file_path, 'r', encoding='utf-8') as f:
                    new_names = [line.strip() for line in f if line.strip()]

            added = 0
            for name in new_names:
                if name not in self.students:
                    self.students.append(name)
                    added += 1

            if added > 0:
                self.students.sort()
                self.save_students()
                self.refresh_student_list()
                self.draw_circle()
                self.info_label.config(text=f"📥 导入 {added} 名学生  |  共 {len(self.students)} 名学生")
                messagebox.showinfo("导入成功", f"成功导入 {added} 名学生！\n（已自动去重）")
            else:
                messagebox.showinfo("提示", "没有新学生需要添加（均已存在）")
        except Exception as e:
            messagebox.showerror("导入失败", f"读取文件出错:\n{e}")

    def export_student_list(self):
        """导出学生名单"""
        file_path = filedialog.asksaveasfilename(
            title="导出学生名单",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("CSV文件", "*.csv")]
        )
        if not file_path:
            return

        try:
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    for name in sorted(self.students):
                        writer.writerow([name])
            else:
                with open(file_path, 'w', encoding='utf-8') as f:
                    for name in sorted(self.students):
                        f.write(name + "\n")
            self.info_label.config(text=f"📋 已导出 {len(self.students)} 名学生")
            messagebox.showinfo("导出成功", f"已导出 {len(self.students)} 名学生到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ================================================================
    #  考勤统计
    # ================================================================
    def record_attendance(self, name, status):
        """记录考勤：present/absent/late"""
        today = datetime.now().strftime("%Y-%m-%d")
        if name not in self.attendance:
            self.attendance[name] = {}
        self.attendance[name][today] = status
        self.save_attendance()
        self.refresh_stats()

    def refresh_stats(self):
        """刷新考勤统计显示"""
        if not hasattr(self, 'stats_label'):
            return

        today = datetime.now().strftime("%Y-%m-%d")
        present = absent = late = 0

        for student, records in self.attendance.items():
            if today in records:
                s = records[today]
                if s == "present":
                    present += 1
                elif s == "absent":
                    absent += 1
                elif s == "late":
                    late += 1

        total = len(self.students)
        unchecked = total - present - absent - late

        text = f"\n"
        text += f"   📋 总人数：{total}\n\n"
        text += f"   ✅ 正常到课：{present}\n"
        text += f"   ❌ 旷课：{absent}\n"
        text += f"   ⏰ 请假：{late}\n"
        text += f"   ⬜ 未记录：{unchecked}\n\n"

        if total > 0:
            recorded = present + absent + late
            rate = (present + late) / total * 100 if total > 0 else 0
            text += f"   📈 出勤率：{rate:.1f}%\n"
            text += f"   📝 已记录：{recorded}/{total}"

        self.stats_label.config(text=text)

    def export_attendance_report(self):
        """导出考勤报表"""
        if not self.attendance:
            messagebox.showinfo("提示", "暂无考勤记录！")
            return

        file_path = filedialog.asksaveasfilename(
            title="导出考勤报表",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("Excel文件", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            today = datetime.now().strftime("%Y-%m-%d")

            if file_path.lower().endswith('.xlsx'):
                wb = Workbook()
                ws = wb.active
                ws.title = "考勤报表"
                ws.append(["姓名", "考勤状态", "日期"])
                for student in sorted(self.students):
                    status_raw = self.attendance.get(student, {}).get(today, "未记录")
                    status_text = {"present": "正常到课", "absent": "旷课",
                                   "late": "请假"}.get(status_raw, "未记录")
                    ws.append([student, status_text, today])
                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 14
                ws.column_dimensions['C'].width = 14
                wb.save(file_path)
            else:
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['姓名', '考勤状态', '日期'])
                    for student in sorted(self.students):
                        status_raw = self.attendance.get(student, {}).get(today, "未记录")
                        status_text = {"present": "正常到课", "absent": "旷课",
                                       "late": "请假"}.get(status_raw, "未记录")
                        writer.writerow([student, status_text, today])

            messagebox.showinfo("导出成功", f"考勤报表已导出到:\n{file_path}")
            self.info_label.config(text=f"📥 考勤报表已导出")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ================================================================
    #  圆形画布 + 动画
    # ================================================================
    def draw_circle(self, highlight_index=-1):
        """在画布上绘制圆形排列的名字"""
        self.canvas.delete("all")
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()
        t = self.theme

        if w < 50 or h < 50:
            w, h = 650, 400

        cx, cy = w // 2, h // 2
        n = len(self.students)
        if n == 0:
            self.canvas.create_text(cx, cy, text="请先添加学生名单",
                                     font=(FONT_FAMILY, 20), fill=t["info_fg"])
            return

        radius = min(cx, cy) - 60

        # 外圈装饰圆环
        self.canvas.create_oval(
            cx - radius, cy - radius, cx + radius, cy + radius,
            outline=t["circle_track"], width=2, dash=(6, 8)
        )
        self.canvas.create_oval(
            cx - radius + 20, cy - radius + 20, cx + radius - 20, cy + radius - 20,
            outline=t["circle_track"], width=1, dash=(3, 6)
        )

        # 轨道上的小点装饰
        for i in range(36):
            angle = (2 * math.pi / 36) * i - math.pi / 2
            dot_x = cx + radius * math.cos(angle)
            dot_y = cy + radius * math.sin(angle)
            self.canvas.create_oval(
                dot_x - 2, dot_y - 2, dot_x + 2, dot_y + 2,
                fill=t["circle_dot"], outline=""
            )

        # 绘制每个学生
        for i, name in enumerate(self.students):
            angle = (2 * math.pi / n) * i - math.pi / 2
            x = cx + radius * math.cos(angle)
            y = cy + radius * math.sin(angle)

            is_highlighted = (i == highlight_index)
            is_selected = (self.selected_student == name and not self.is_spinning)

            if is_highlighted:
                # 光晕效果
                glow_sizes = [42, 36, 28]
                glow_colors = ["#fed330", "#f7b731", "#f39c12"]
                for gs, gc in zip(glow_sizes, glow_colors):
                    self.canvas.create_oval(
                        x - gs, y - gs, x + gs, y + gs,
                        fill=gc, outline="", stipple="gray50"
                    )
                self.canvas.create_oval(
                    x - 20, y - 20, x + 20, y + 20,
                    fill=t["highlight"], outline="#f39c12", width=3
                )
                self.canvas.create_text(
                    x, y + 32, text=name,
                    font=(FONT_FAMILY, 13, "bold"),
                    fill=t["highlight"]
                )
            elif is_selected:
                self.canvas.create_oval(
                    x - 18, y - 18, x + 18, y + 18,
                    fill=t["selected"], outline="#eb3b5a", width=2
                )
                self.canvas.create_text(
                    x, y + 30, text=name,
                    font=(FONT_FAMILY, 13, "bold"),
                    fill=t["selected"]
                )
            else:
                self.canvas.create_oval(
                    x - 8, y - 8, x + 8, y + 8,
                    fill=t["circle_dot"], outline=t["circle_dot_outline"], width=1
                )
                self.canvas.create_text(
                    x, y + 22, text=name,
                    font=(FONT_FAMILY, 10),
                    fill=t["circle_name"]
                )

        # 中心文字
        if self.is_spinning:
            center_text = "旋转中..."
            center_color = t["highlight"]
        elif self.selected_student:
            center_text = f"🎉 {self.selected_student}"
            center_color = t["selected"]
        else:
            center_text = "点击「开始点名」\n或按 F5"
            center_color = t["info_fg"]

        self.canvas.create_text(
            cx, cy, text=center_text,
            font=(FONT_FAMILY, 16, "bold"),
            fill=center_color
        )

    def start_spin(self):
        if self.is_spinning:
            return
        if len(self.students) == 0:
            messagebox.showwarning("提示", "请先添加学生名单！")
            return

        # 检查冷却时间
        if hasattr(self, '_last_call_time'):
            elapsed = (time.time() - self._last_call_time) * 1000
            if elapsed < self.call_cooldown_ms:
                return

        self._last_call_time = time.time()
        self.is_spinning = True
        self.selected_student = None
        self.spin_btn.config(state=tk.DISABLED, text="⏳ 旋转中...")
        self.info_label.config(text="灯光旋转中...")
        self.spin_start_time = time.time()
        self.spin_speed = 0.03
        self.current_index = 0
        self.animate_spin()

    def animate_spin(self):
        if not self.is_spinning:
            return

        elapsed = time.time() - self.spin_start_time
        total_duration = self.spin_duration

        if elapsed >= total_duration:
            self.is_spinning = False
            self.selected_student = self.students[self.current_index % len(self.students)]
            self.draw_circle(highlight_index=-1)
            self.spin_btn.config(state=tk.NORMAL, text="🎰 开始点名 (F5)")
            self.info_label.config(text=f"选中: {self.selected_student}  |  共 {len(self.students)} 名学生")

            # 语音播报
            self.speak_name(self.selected_student)

            # 弹出评分对话框
            self.root.after(300, self.show_scoring_dialog)
            return

        progress = elapsed / total_duration

        if progress < 0.6:
            speed = 0.03 + 0.02 * math.sin(progress * 8)
        elif progress < 0.85:
            speed = 0.06 + 0.04 * (progress - 0.6) / 0.25
        else:
            remaining = 1.0 - progress
            speed = 0.10 + 0.30 * (remaining / 0.15) ** 2

        if progress > 0.88 and random.random() < 0.3:
            speed = speed * 3

        self.spin_speed = speed
        self.current_index = (self.current_index + 1) % len(self.students)
        self.draw_circle(highlight_index=self.current_index)

        delay_ms = int(speed * 1000)
        self.root.after(max(delay_ms, 20), self.animate_spin)

    def re_speak(self):
        """重新播报当前选中的学生"""
        if self.selected_student and not self.is_spinning:
            self.speak_name(self.selected_student)
            self.info_label.config(text=f"🔊 重新播报: {self.selected_student}")
        elif self.is_spinning:
            pass  # 旋转中不播报
        else:
            messagebox.showinfo("提示", "请先进行点名！")

    # ================================================================
    #  评分对话框（含考勤）
    # ================================================================
    def show_scoring_dialog(self):
        if not self.selected_student:
            return

        name = self.selected_student
        t = self.theme

        dialog = tk.Toplevel(self.root)
        dialog.title(f"📝 记录回答 - {name}")
        dialog.geometry("520x480")
        dialog.configure(bg=t["panel_bg"])
        dialog.transient(self.root)
        dialog.grab_set()

        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() - w) // 2
        y = (dialog.winfo_screenheight() - h) // 2
        dialog.geometry(f"+{x}+{y}")

        # 标题
        tk.Label(
            dialog, text=f"🎤 {name} 同学请回答",
            font=(FONT_FAMILY, 16, "bold"),
            fg=t["highlight"], bg=t["panel_bg"]
        ).pack(pady=(20, 10))

        # 回答输入区
        input_frame = tk.Frame(dialog, bg=t["panel_bg"])
        input_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=5)

        tk.Label(
            input_frame, text="📝 回答内容：",
            font=(FONT_FAMILY, 11),
            fg=t["text"], bg=t["panel_bg"], anchor=tk.W
        ).pack(anchor=tk.W)

        answer_text = tk.Text(
            input_frame, font=(FONT_FAMILY, 11),
            height=4, wrap=tk.WORD,
            bg=t["input_bg"], fg=t["input_fg"],
            insertbackground=t["text"] if self._dark_mode else "black",
            relief=tk.FLAT, padx=10, pady=8,
            highlightthickness=1, highlightbackground=t["accent"]
        )
        answer_text.pack(fill=tk.BOTH, expand=True, pady=(5, 10))

        # 评分区
        score_frame = tk.Frame(dialog, bg=t["panel_bg"])
        score_frame.pack(fill=tk.X, padx=25, pady=(0, 8))

        tk.Label(
            score_frame, text="⭐ 评分 (1-10)：",
            font=(FONT_FAMILY, 11),
            fg=t["text"], bg=t["panel_bg"]
        ).pack(side=tk.LEFT)

        score_var = tk.IntVar(value=8)
        score_scale = tk.Scale(
            score_frame, from_=1, to=10, orient=tk.HORIZONTAL,
            variable=score_var, length=180,
            bg=t["panel_bg"], fg=t["highlight"],
            troughcolor=t["input_bg"],
            highlightthickness=0,
            font=(FONT_FAMILY, 10)
        )
        score_scale.pack(side=tk.RIGHT, padx=10)

        # 快捷评分按钮
        quick_frame = tk.Frame(dialog, bg=t["panel_bg"])
        quick_frame.pack(fill=tk.X, padx=25, pady=(0, 5))

        for s in [2, 4, 6, 8, 10]:
            btn = tk.Button(
                quick_frame, text=f"{s}分", font=(FONT_FAMILY, 9),
                bg=t["score_colors"][(s // 2) - 1], fg="white",
                activebackground=t["score_colors"][(s // 2) - 1],
                relief=tk.FLAT, padx=10, cursor="hand2",
                command=lambda val=s: score_var.set(val)
            )
            btn.pack(side=tk.LEFT, padx=3)

        # ---- 考勤按钮区（新增） ----
        att_frame = tk.Frame(dialog, bg=t["panel_bg"])
        att_frame.pack(fill=tk.X, padx=25, pady=(8, 5))

        tk.Label(
            att_frame, text="📋 考勤：",
            font=(FONT_FAMILY, 11),
            fg=t["text"], bg=t["panel_bg"]
        ).pack(side=tk.LEFT, padx=(0, 8))

        att_var = tk.StringVar(value="")  # 空 = 不记录

        tk.Button(
            att_frame, text="✅ 正常到课", font=(FONT_FAMILY, 9),
            bg=t["btn_save"], fg="white",
            activebackground=t["btn_save_hover"],
            relief=tk.FLAT, padx=8, pady=3, cursor="hand2",
            command=lambda: att_var.set("present")
        ).pack(side=tk.LEFT, padx=2)

        tk.Button(
            att_frame, text="❌ 旷课", font=(FONT_FAMILY, 9),
            bg=t["btn_danger"], fg="white",
            activebackground=t["btn_danger_hover"],
            relief=tk.FLAT, padx=8, pady=3, cursor="hand2",
            command=lambda: att_var.set("absent")
        ).pack(side=tk.LEFT, padx=2)

        tk.Button(
            att_frame, text="⏰ 请假", font=(FONT_FAMILY, 9),
            bg=t["btn_warning"], fg="white",
            activebackground=t["btn_warning_hover"],
            relief=tk.FLAT, padx=8, pady=3, cursor="hand2",
            command=lambda: att_var.set("late")
        ).pack(side=tk.LEFT, padx=2)

        att_status_label = tk.Label(
            att_frame, text="", font=(FONT_FAMILY, 9),
            bg=t["panel_bg"], fg=t["highlight"]
        )
        att_status_label.pack(side=tk.LEFT, padx=8)

        # 跟踪 att_var 变化更新标签
        def on_att_change(*args):
            status_map = {"present": "✅ 正常到课", "absent": "❌ 旷课", "late": "⏰ 请假"}
            att_status_label.config(text=status_map.get(att_var.get(), ""))
        att_var.trace('w', on_att_change)

        # 按钮行
        btn_frame = tk.Frame(dialog, bg=t["panel_bg"])
        btn_frame.pack(pady=(10, 20))

        def on_save():
            answer = answer_text.get("1.0", tk.END).strip()
            if not answer:
                messagebox.showwarning("提示", "请输入回答内容！", parent=dialog)
                return

            score = score_var.get()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            time_short = datetime.now().strftime("%H:%M:%S")

            # 保存到 Excel
            success = self.save_to_excel(name, answer, score, now)

            # 记录考勤
            att_status = att_var.get()
            if att_status:
                self.record_attendance(name, att_status)

            # 添加到历史
            hist_entry = f"[{time_short}] 🎲 {name} → 评分 {score}/10"
            if att_status:
                status_txt = {"present": "✅到课", "absent": "❌旷课", "late": "⏰请假"}[att_status]
                hist_entry += f"  {status_txt}"
            self.add_history(hist_entry)

            if success:
                msg = f"已记录 {name} 的回答！\n评分: {score}/10"
                if att_status:
                    status_txt = {"present": "正常到课", "absent": "旷课", "late": "请假"}[att_status]
                    msg += f"\n考勤: {status_txt}"
                self.info_label.config(
                    text=f"✅ 已记录: {name} - {score}分  |  Excel: {EXCEL_FILE.name}"
                )
                messagebox.showinfo("成功", msg, parent=dialog)
            else:
                messagebox.showerror("错误", "保存失败，请检查文件权限", parent=dialog)

            dialog.destroy()

        tk.Button(
            btn_frame, text="💾 保存记录", font=(FONT_FAMILY, 12, "bold"),
            bg=t["btn_save"], fg="white",
            activebackground=t["btn_save_hover"],
            relief=tk.FLAT, padx=25, pady=6, cursor="hand2",
            command=on_save
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame, text="❌ 跳过", font=(FONT_FAMILY, 12),
            bg=t["btn_cancel"], fg="white",
            activebackground=t["btn_cancel_hover"],
            relief=tk.FLAT, padx=25, pady=6, cursor="hand2",
            command=dialog.destroy
        ).pack(side=tk.LEFT, padx=8)

        answer_text.focus_set()

    # ================================================================
    #  学生管理对话框（旧版保留，用于编辑全部名单）
    # ================================================================
    def open_student_manager(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("📋 管理学生名单")
        dialog.geometry("500x550")
        dialog.configure(bg=self.theme["panel_bg"])
        dialog.transient(self.root)
        dialog.grab_set()

        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() - w) // 2
        y = (dialog.winfo_screenheight() - h) // 2
        dialog.geometry(f"+{x}+{y}")

        t = self.theme

        tk.Label(
            dialog, text="📋 学生名单管理",
            font=(FONT_FAMILY, 16, "bold"),
            fg=t["highlight"], bg=t["panel_bg"]
        ).pack(pady=(20, 10))

        tk.Label(
            dialog, text="每行一个学生姓名，编辑后点击保存即可。",
            font=(FONT_FAMILY, 9),
            fg=t["info_fg"], bg=t["panel_bg"]
        ).pack(pady=(0, 10))

        text_frame = tk.Frame(dialog, bg=t["panel_bg"])
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        student_text = tk.Text(
            text_frame, font=(FONT_FAMILY, 12),
            wrap=tk.WORD,
            bg=t["input_bg"], fg=t["input_fg"],
            insertbackground=t["text"] if self._dark_mode else "black",
            relief=tk.FLAT, padx=12, pady=10,
            highlightthickness=1, highlightbackground=t["accent"]
        )
        student_text.pack(fill=tk.BOTH, expand=True)
        student_text.insert("1.0", "\n".join(self.students))

        btn_frame = tk.Frame(dialog, bg=t["panel_bg"])
        btn_frame.pack(pady=(12, 20))

        def on_save():
            content = student_text.get("1.0", tk.END).strip()
            names = [n.strip() for n in content.split("\n") if n.strip()]
            if not names:
                messagebox.showwarning("提示", "至少需要一个学生！", parent=dialog)
                return
            self.students = names
            self.save_students()
            self.selected_student = None
            self.refresh_student_list()
            self.draw_circle()
            self.refresh_stats()
            self.info_label.config(text=f"共 {len(self.students)} 名学生  |  名单已更新")
            messagebox.showinfo("成功", f"已保存 {len(names)} 名学生！", parent=dialog)
            dialog.destroy()

        tk.Button(
            btn_frame, text="💾 保存名单", font=(FONT_FAMILY, 12, "bold"),
            bg=t["btn_save"], fg="white",
            activebackground=t["btn_save_hover"],
            relief=tk.FLAT, padx=25, pady=6, cursor="hand2",
            command=on_save
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame, text="↩ 恢复默认", font=(FONT_FAMILY, 12),
            bg=t["btn_cancel"], fg="white",
            activebackground=t["btn_cancel_hover"],
            relief=tk.FLAT, padx=18, pady=6, cursor="hand2",
            command=lambda: [student_text.delete("1.0", tk.END),
                            student_text.insert("1.0", "\n".join(DEFAULT_STUDENTS))]
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame, text="❌ 取消", font=(FONT_FAMILY, 12),
            bg="#2d3436", fg="white",
            activebackground="#636e72",
            relief=tk.FLAT, padx=18, pady=6, cursor="hand2",
            command=dialog.destroy
        ).pack(side=tk.LEFT, padx=8)

    # ---- 查看记录 ----
    def open_records(self):
        if EXCEL_FILE.exists():
            try:
                os.startfile(str(EXCEL_FILE))
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件:\n{e}")
        else:
            messagebox.showinfo("提示", "还没有记录文件。\n开始点名并保存记录后会自动生成。")

    # ---- 历史刷新 ----
    def refresh_history_tab(self):
        if not hasattr(self, 'history_text'):
            return
        self.history_text.config(state=tk.NORMAL)
        self.history_text.delete("1.0", tk.END)
        for record in reversed(self.call_history[-30:]):
            self.history_text.insert(tk.END, record + "\n")
        self.history_text.config(state=tk.DISABLED)
        self.history_text.see("1.0")

    def on_close(self):
        self.save_history()
        self.save_attendance()
        self.root.destroy()


def main():
    root = tk.Tk()

    # 高DPI支持
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = RollCallApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
